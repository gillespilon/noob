#! /usr/bin/env python3
"""
- Create three dataframes
- Write one dataframe to a csv file
- Read a csv file and correct dtypes
- Write three dataframes to three worksheets to an Excel workbook
- Read an Excel workbook with three worksheets and correct dtypes
- Read an Excel workbook with data and formulae in three worksheets

./02_data_file_write.py
"""

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl import Workbook
import datasense as ds
import pandas as pd


pd.options.display.width = 160
output_url = '02_data_file_write.html'
header_title = 'Writing a data file'
header_id = 'writing-data-file'


def main():
    original_stdout = ds.html_begin(
        output_url=output_url,
        header_title=header_title,
        header_id=header_id
    )
    print('Create three dataframes')
    print()
    size = 4
    df1 = ds.create_dataframe(size=size)
    df2 = ds.create_dataframe(size=size)
    df3 = ds.create_dataframe(size=size)
    print('df1')
    print(df1.head(2))
    print()
    print(df1.dtypes)
    print()
    print('df2')
    print(df2.head(2))
    print()
    print(df2.dtypes)
    print()
    print('df3')
    print(df3.head(2))
    print()
    print(df3.dtypes)
    print()
    print('Write one dataframe to a csv file')
    print()
    ds.save_file(
        df=df1,
        file_name='data/just_a_test.csv'
    )
    print('Saved csv file to data/just_a_test.csv')
    print()
    print('Read a csv file and do not correct dtypes')
    print()
    just_a_test = ds.read_file(
        file_name='data/just_a_test.csv'
    )
    print(just_a_test.head(2))
    print()
    print(just_a_test.dtypes)
    print()
    print('Read a csv file and correct dtypes')
    print()
    convert_dict = {
        'a': 'float64',
        'b': 'boolean',
        'c': 'category',
        'i': 'float64',
        'r': 'str',
        's': 'str',
        'x': 'float64',
        'y': 'Int64',
        'z': 'float64'
    }
    parse_dates = ['t', 'u']
    time_delta_columns = ['d']
    just_a_test = ds.read_file(
        file_name='data/just_a_test.csv',
        dtype=convert_dict,
        parse_dates=parse_dates,
        time_delta_columns=time_delta_columns
    )
    print(just_a_test.head(2))
    print()
    print(just_a_test.dtypes)
    print()
    print('Write three dataframes to three worksheets to an Excel workbook')
    print()
    path = 'data/even_another_file.xlsx'
    engine = 'openpyxl'
    with pd.ExcelWriter(path=path, engine=engine) as writer:
        df1.to_excel(
            excel_writer=writer,
            sheet_name='sheet_one',
            index=False
        )
        df2.to_excel(
            excel_writer=writer,
            sheet_name='sheet_two',
            index=False
        )
        df3.to_excel(
            excel_writer=writer,
            sheet_name='sheet_three',
            index=False
        )
    writer.save()
    print('Saved xlsx file to data/event_another_file.xlsx')
    print()
    print('Read an Excel workbook with three worksheets and correct dtypes')
    print()
    print('openpyxl')
    print()
    wb1 = load_workbook(filename='data/even_another_file.xlsx')
    print(wb1.sheetnames)
    print()
    wb1ws1 = wb1['sheet_one']
    from itertools import islice
    data = wb1ws1.values
    cols = next(data)[:]
    data = list(data)
    # idx = [row[0] for row in data]
    data = (islice(row, None) for row in data)
    wb1df1 = pd.DataFrame(data, columns=cols)
    print(wb1df1.head(2))
    print()
    print(wb1df1.dtypes)
    print()
    wb1df1 = wb1df1.astype(convert_dict)
    for column in time_delta_columns:
        wb1df1[column] = pd.to_timedelta(wb1df1[column])
    wb1df1[parse_dates] = wb1df1[parse_dates].astype(dtype='datetime64[ns]')
    print(wb1df1.dtypes)
    print()
    wb1ws2 = wb1['sheet_two']
    from itertools import islice
    data = wb1ws2.values
    cols = next(data)[:]
    data = list(data)
    # idx = [row[0] for row in data]
    data = (islice(row, None) for row in data)
    wb1df2 = pd.DataFrame(data, columns=cols)
    print(wb1df2.head(2))
    print()
    print(wb1df2.dtypes)
    print()
    wb1df2 = wb1df2.astype(convert_dict)
    for column in time_delta_columns:
        wb1df2[column] = pd.to_timedelta(wb1df2[column])
    wb1df2[parse_dates] = wb1df2[parse_dates].astype(dtype='datetime64[ns]')
    print(wb1df2.dtypes)
    print()
    wb1ws3 = wb1['sheet_three']
    from itertools import islice
    data = wb1ws3.values
    cols = next(data)[:]
    data = list(data)
    # idx = [row[0] for row in data]
    data = (islice(row, None) for row in data)
    wb1df3 = pd.DataFrame(data, columns=cols)
    print(wb1df3.head(2))
    print()
    print(wb1df3.dtypes)
    print()
    wb1df3 = wb1df3.astype(convert_dict)
    for column in time_delta_columns:
        wb1df3[column] = pd.to_timedelta(wb1df3[column])
    wb1df3[parse_dates] = wb1df3[parse_dates].astype(dtype='datetime64[ns]')
    print(wb1df3.dtypes)
    print()
    print('pd.read_excel')
    print()
    wb2 = pd.read_excel(
        io=path,
        sheet_name=None,
        engine='openpyxl'
    )
    print(wb2)
    print()
    print(wb2.keys())
    print()
    wb2df1 = wb2['sheet_one']
    print(wb2df1.head(2))
    print()
    print(wb2df1.dtypes)
    print()
    wb2df1 = wb2df1.astype(convert_dict)
    for column in time_delta_columns:
        wb2df1[column] = pd.to_timedelta(wb2df1[column])
    wb2df1[parse_dates] = wb2df1[parse_dates].astype(dtype='datetime64[ns]')
    print(wb2df1.dtypes)
    print()
    wb2df2 = wb2['sheet_two']
    print(wb2df2.head(2))
    print()
    print(wb2df2.dtypes)
    print()
    wb2df2 = wb2df1.astype(convert_dict)
    for column in time_delta_columns:
        wb2df2[column] = pd.to_timedelta(wb2df2[column])
    wb2df2[parse_dates] = wb2df2[parse_dates].astype(dtype='datetime64[ns]')
    print(wb2df2.dtypes)
    print()
    wb2df3 = wb2['sheet_three']
    print(wb2df3.head(2))
    print()
    print(wb2df3.dtypes)
    print()
    wb2df3 = wb2df3.astype(convert_dict)
    for column in time_delta_columns:
        wb2df3[column] = pd.to_timedelta(wb2df3[column])
    wb2df3[parse_dates] = wb2df3[parse_dates].astype(dtype='datetime64[ns]')
    print(wb2df3.dtypes)
    print()
    print('Read an Excel workbook with data and formulae in three worksheets')
    print()
    print('openpyxl')
    print()
    wb3 = load_workbook(filename='data/file_with_formulae.xlsx')
    print(wb3.sheetnames)
    print()
    print('sheet_calcs_1')
    wb3s1 = wb3['sheet_calcs_1']
    wb3s1['C1'] = 'square_root'
    for row in range(2, 9):
        wb3s1[f'C{row}'] = f'=B{row}^0.5'
    from itertools import islice
    data = wb3s1.values
    cols = next(data)[:]
    data = list(data)
    # idx = [row[0] for row in data]
    data = (islice(row, None) for row in data)
    wb3df1 = pd.DataFrame(data, columns=cols)
    print(wb3df1)
    print()
    print(wb3df1.dtypes)
    print()
    print('sheet_calcs_2')
    wb3s2 = wb3['sheet_calcs_2']
    wb3s2['C1'] = 'cube_root'
    for row in range(2, 9):
        wb3s2[f'C{row}'] = f'=B{row}^(1/3)'
    from itertools import islice
    data = wb3s2.values
    cols = next(data)[:]
    data = list(data)
    # idx = [row[0] for row in data]
    data = (islice(row, None) for row in data)
    wb3df2 = pd.DataFrame(data, columns=cols)
    print(wb3df2)
    print()
    print(wb3df2.dtypes)
    print()
    print('sheet_calcs_3')
    wb3s3 = wb3['sheet_calcs_3']
    wb3s3['C1'] = 'fourth_root'
    for row in range(2, 9):
        wb3s3[f'C{row}'] = f'=B{row}^(1/4)'
    from itertools import islice
    data = wb3s3.values
    cols = next(data)[:]
    data = list(data)
    # idx = [row[0] for row in data]
    data = (islice(row, None) for row in data)
    wb3df3 = pd.DataFrame(data, columns=cols)
    print(wb3df3)
    print()
    print(wb3df3.dtypes)
    print()
    path = 'data/file_with_formulae_plus.xlsx'
    engine = 'openpyxl'
    with pd.ExcelWriter(path=path, engine=engine) as writer:
        wb3df1.to_excel(
            excel_writer=writer,
            sheet_name='sheet_one',
            index=False
        )
        wb3df2.to_excel(
            excel_writer=writer,
            sheet_name='sheet_two',
            index=False
        )
        wb3df3.to_excel(
            excel_writer=writer,
            sheet_name='sheet_three',
            index=False
        )
    writer.save()
    print('pd.read_excel')
    print()
    wb4 = pd.read_excel(
        io='data/file_with_formulae.xlsx',
        sheet_name=None,
        engine='openpyxl'
    )
    print(wb4.keys())
    print()
    print('sheet_calcs_1')
    wb4s1 = wb4['sheet_calcs_1']
    print(wb4s1)
    print()
    print(wb4s1.dtypes)
    print()
    print('sheet_calcs_2')
    wb4s2 = wb4['sheet_calcs_2']
    print(wb4s2)
    print()
    print(wb4s2.dtypes)
    print()
    print('sheet_calcs_3')
    wb4s3 = wb4['sheet_calcs_3']
    print(wb4s3)
    print()
    print(wb4s3.dtypes)
    ds.html_end(
        original_stdout=original_stdout,
        output_url=output_url
    )


if __name__ == '__main__':
    main()
